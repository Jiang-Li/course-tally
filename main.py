import os
from pathlib import Path
import pandas as pd
from src.tally_reader import read_tally_file
from src.excel_handler import find_leeds_file, read_leeds_courses, ExcelFileError
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import re

def analyze_duplicates(df: pd.DataFrame) -> tuple[bool, pd.DataFrame]:
    """
    Analyze duplicates in a DataFrame.
    
    Args:
        df (pd.DataFrame): DataFrame to analyze
        
    Returns:
        tuple[bool, pd.DataFrame]: (has_duplicates, duplicate_rows)
    """
    duplicates = df.duplicated()
    duplicate_rows = df[df.duplicated(keep=False)] if duplicates.sum() > 0 else pd.DataFrame()
    return bool(duplicates.sum()), duplicate_rows

def clean_column_name(col_name: str) -> str:
    """
    Clean column name by removing spaces, special characters, newlines, and tabs.
    
    Args:
        col_name (str): Original column name
        
    Returns:
        str: Cleaned column name
    """
    # Convert to string in case it's not
    col_name = str(col_name)
    # Remove newlines and tabs first
    col_name = col_name.replace('\n', '').replace('\t', '').strip()
    # Remove special characters and spaces, keep only alphanumeric
    cleaned = re.sub(r'[^a-zA-Z0-9]', '', col_name)
    # Convert to lowercase for consistent comparison
    return cleaned.lower()

def display_column_mapping(leeds_df: pd.DataFrame, tally_df: pd.DataFrame) -> None:
    """
    Display a table showing matched and unmatched columns between Leeds and Tally files.
    
    Args:
        leeds_df (pd.DataFrame): Leeds DataFrame
        tally_df (pd.DataFrame): Tally DataFrame
    """
    # Get cleaned column names
    leeds_cols = {clean_column_name(col): col for col in leeds_df.columns}
    tally_cols = {clean_column_name(col): col for col in tally_df.columns}
    
    # Create lists for the table
    all_cleaned_cols = sorted(set(leeds_cols.keys()) | set(tally_cols.keys()))
    rows = []
    
    for cleaned_col in all_cleaned_cols:
        leeds_original = leeds_cols.get(cleaned_col, '')
        tally_original = tally_cols.get(cleaned_col, '')
        status = 'Matched' if cleaned_col in leeds_cols and cleaned_col in tally_cols else 'Unmatched'
        rows.append([cleaned_col, leeds_original, tally_original, status])
    
    # Create DataFrame for display
    mapping_df = pd.DataFrame(rows, columns=['Cleaned Name', 'Leeds Original', 'Tally Original', 'Status'])
    print("\nColumn Mapping Table:")
    print(mapping_df.to_string(index=False))
    
    # Print summary
    matched = mapping_df[mapping_df['Status'] == 'Matched']
    unmatched = mapping_df[mapping_df['Status'] == 'Unmatched']
    print(f"\nSummary:")
    print(f"  Matched columns: {len(matched)}")
    print(f"  Unmatched columns: {len(unmatched)}")

def clean_course_number(val) -> int:
    """
    Convert course number directly to integer.
    
    Args:
        val: Course number value to convert
        
    Returns:
        int: Course number as integer
    """
    try:
        return int(val)
    except:
        return 0

def clean_days(val) -> str:
    """
    Clean days value by removing spaces.
    
    Args:
        val: Days value to clean
        
    Returns:
        str: Days value without spaces
    """
    try:
        # Convert to string and remove all spaces
        return str(val).replace(' ', '')
    except:
        return ''

def compare_and_update_courses(tally_df: pd.DataFrame, leeds_df: pd.DataFrame, leeds_file: str) -> None:
    """
    Compare tally and Leeds course data, report unmatched courses, and update Leeds data with Tally values.
    
    Args:
        tally_df (pd.DataFrame): DataFrame from tally file
        leeds_df (pd.DataFrame): DataFrame from Leeds file
        leeds_file (str): Path to Leeds file for reference
    """
    # Clean column names in both DataFrames
    leeds_df = leeds_df.copy()
    tally_df = tally_df.copy()
    
    # Create mapping of original to cleaned names
    leeds_col_map = {clean_column_name(col): col for col in leeds_df.columns}
    tally_col_map = {clean_column_name(col): col for col in tally_df.columns}
    
    # Rename columns with cleaned names
    leeds_df.columns = [clean_column_name(col) for col in leeds_df.columns]
    tally_df.columns = [clean_column_name(col) for col in tally_df.columns]
    
    # Convert course numbers to integers in both DataFrames
    if 'crsno' in leeds_df.columns:
        leeds_df['crsno'] = leeds_df['crsno'].apply(clean_course_number)
        
    if 'crsno' in tally_df.columns:
        tally_df['crsno'] = tally_df['crsno'].apply(clean_course_number)
    
    # Clean days column in both DataFrames
    if 'days' in leeds_df.columns:
        leeds_df['days'] = leeds_df['days'].apply(clean_days)
        
    if 'days' in tally_df.columns:
        tally_df['days'] = tally_df['days'].apply(clean_days)
    
    # Define the key columns for matching rows (using cleaned names)
    key_cols = ['subj', 'crsno', 'sec', 'days']
    
    # Track unmatched rows and updates
    unmatched_leeds = leeds_df.copy()
    updates_made = False
    
    # Load the Excel workbook to preserve formatting
    wb = load_workbook(leeds_file)
    ws = wb.active
    
    # Find the header row
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and clean_column_name(cell.value) == 'subj':
                header_row = cell.row
                break
        if header_row:
            break
    
    if not header_row:
        print("Error: Could not find header row in Leeds file")
        return
        
    # Create column index mapping
    col_indices = {}
    for col in ws[header_row]:
        if col.value:
            cleaned_name = clean_column_name(col.value)
            col_indices[cleaned_name] = col.column
    
    # Find matches and update values
    for _, tally_row in tally_df.iterrows():
        # Create mask for matching rows in Leeds data
        mask = True
        for col in key_cols:
            mask = mask & (leeds_df[col] == tally_row[col])
        
        matching_leeds = leeds_df[mask]
        if not matching_leeds.empty:
            unmatched_leeds = unmatched_leeds[~mask]
            
            # Check for differences in non-key columns and update if needed
            leeds_idx = matching_leeds.index[0]
            for col in set(leeds_df.columns) & set(tally_df.columns) - set(key_cols):
                leeds_val = str(matching_leeds.iloc[0][col])
                tally_val = str(tally_row[col])
                
                # Compare values, treating empty strings and NaN as equal
                leeds_val = '' if pd.isna(leeds_val) or leeds_val == 'nan' else leeds_val
                tally_val = '' if pd.isna(tally_val) or tally_val == 'nan' else tally_val
                
                if leeds_val != tally_val and col in col_indices:
                    # Update the Excel cell while preserving formatting
                    excel_row = header_row + leeds_idx + 1
                    excel_col = col_indices[col]
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.value = tally_row[col]
                    updates_made = True
    
    # Save the workbook if updates were made
    if updates_made:
        wb.save(leeds_file)
        print("\nUpdated Leeds file with values from Tally")
    
    # Print results
    if unmatched_leeds.empty:
        print("\nAll Leeds courses match with Tally.")
    else:
        print("\nUnmatched courses in Leeds:")
        unmatched_display = unmatched_leeds[key_cols].rename(columns=leeds_col_map)
        print(unmatched_display.to_string())

def main():
    try:
        # Read tally file
        tally_file = "Course Tally.xlsx"
        table, _, _ = read_tally_file(tally_file)
        
        # Check for duplicates
        has_duplicates, duplicate_rows = analyze_duplicates(table)
        if has_duplicates:
            print(f"Found {len(duplicate_rows)} duplicate rows:")
            print(duplicate_rows)
            
        # Find and read Leeds courses file
        try:
            leeds_file = find_leeds_file(tally_file)
            leeds_table = read_leeds_courses(leeds_file)
            
            # Compare and update courses
            compare_and_update_courses(table, leeds_table, leeds_file)
            
        except ExcelFileError as e:
            print(f"Error processing Leeds file: {str(e)}")
            return

    except Exception as e:
        print(f"Error: {str(e)}")
        return

if __name__ == "__main__":
    main() 